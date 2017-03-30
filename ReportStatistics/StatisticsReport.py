# -*- coding: utf-8 -*-
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from pprint import pprint


def read_report(file_name):
    with open(file_name, encoding='UTF-8') as f:
        report_content = f.read()
    soup = BeautifulSoup(report_content, 'lxml')
    page_tag_list = soup.select('tr')

    total_statistics_dict = dict()
    c_statistics_dict = dict()
    h_statistics_dict = dict()
    b_statistics_dict = dict()
    for tag in page_tag_list:
        if 'class' in tag.attrs:
            if tag.attrs['class'] == ['hiddenRow'] or tag.attrs['class'] == ['none']:
                status_flag = dict()
                if tag.a.string.strip() == 'pass':
                    status_flag = {'pass': 1, 'error': 0, 'fail': 0}
                if tag.a.string.strip() == 'error':
                    status_flag = {'pass': 0, 'error': 1, 'fail': 0}
                if tag.a.string.strip() == 'fail':
                    status_flag = {'pass': 0, 'error': 0, 'fail': 1}

                # 注意!tag.td.div.string的类型为NavigableString

                if tag.td.div.string.startswith('C端'):
                    c_statistics_dict[tag.td.div.string] = status_flag
                if tag.td.div.string.startswith('H端'):
                    h_statistics_dict[tag.td.div.string] = status_flag
                if tag.td.div.string.startswith('B端'):
                    b_statistics_dict[tag.td.div.string] = status_flag

    total_statistics_dict['C端统计'] = c_statistics_dict
    total_statistics_dict['H端统计'] = h_statistics_dict
    total_statistics_dict['B端统计'] = b_statistics_dict
    # pprint(total_statistics_dict)
    return total_statistics_dict


def read_xlsx():
    statistics_file = os.path.abspath('../ReportAndLog/Statistics/Online_System_Status_Statistics.xlsx')
    wb = load_workbook(statistics_file)
    workbook_dict = dict()
    for sheet_name in wb.get_sheet_names():
        if sheet_name == 'C端统计':
            sheet = wb.get_sheet_by_name('C端统计')
            workbook_dict[sheet_name] = dict()
            read_xlsx_action(sheet, workbook_dict[sheet_name])
        if sheet_name == 'H端统计':
            sheet = wb.get_sheet_by_name('H端统计')
            workbook_dict[sheet_name] = dict()
            read_xlsx_action(sheet, workbook_dict[sheet_name])
        if sheet_name == 'B端统计':
            sheet = wb.get_sheet_by_name('B端统计')
            workbook_dict[sheet_name] = dict()
            read_xlsx_action(sheet, workbook_dict[sheet_name])
    # pprint(workbook_dict)
    return workbook_dict


def read_xlsx_action(sheet, sheet_dict):
    for line_no in range(2, sheet.max_row + 1):
        tag_title = sheet['A' + str(line_no)].value
        tag_pass = sheet['B' + str(line_no)].value
        tag_error = sheet['C' + str(line_no)].value
        tag_fail = sheet['D' + str(line_no)].value
        sheet_dict[tag_title] = {'pass': tag_pass, 'error': tag_error, 'fail': tag_fail}


def merge_dict_data(report_dict, file_dict):
    '''
    :description: 合并报告和以前的错误统计
    :param report_dict:
    :param file_dict:
    :return:
    '''
    for report_data_key, report_data_value in report_dict.items():
        if report_data_key in file_dict.keys():
            merge_subject_data(report_data_value, file_dict[report_data_key])
    # pprint(report_dict)


def merge_subject_data(report_data, file_data):
    for report_tag_key, report_statistics in report_data.items():
        if report_tag_key in file_data.keys():
            merge_statistics_data(report_statistics, file_data[report_tag_key])


def merge_statistics_data(report_statistics_data, file_statistics_data):
    for report_stat_key, report_stat_value in report_statistics_data.items():
        if report_stat_key in file_statistics_data.keys():
            report_statistics_data[report_stat_key] = report_stat_value + file_statistics_data[report_stat_key]


def write_xlsx(t_stat_dict):
    statistics_file = os.path.abspath('../ReportAndLog/Statistics/Online_System_Status_Statistics.xlsx')
    # 删除原来的统计文件
    if os.path.exists(statistics_file):
        os.remove(statistics_file)
    wb = Workbook()

    for title, data in t_stat_dict.items():
        if title == 'C端统计':
            sheet = wb.create_sheet(index=0, title='C端统计')
            write_xlsx_action(sheet, data)
        if title == 'H端统计':
            sheet = wb.create_sheet(index=1, title='H端统计')
            write_xlsx_action(sheet, data)
        if title == 'B端统计':
            sheet = wb.create_sheet(index=2, title='B端统计')
            write_xlsx_action(sheet, data)
    # 删除空白的Sheet
    del_sheet = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(del_sheet)

    wb.save(statistics_file)


def write_xlsx_action(sheet, data):
    line_no = 2
    for tag, num_dict in data.items():
        sheet['A' + str(line_no)] = tag
        sheet['B' + str(line_no)] = num_dict['pass']
        sheet['C' + str(line_no)] = num_dict['error']
        sheet['D' + str(line_no)] = num_dict['fail']
        # print(tag, num_dict)
        line_no += 1
    sheet['B1'] = 'Pass'
    sheet['C1'] = 'Error'
    sheet['D1'] = 'Fail'


def run_statistics_report(report_path_name):
    '''
    :description: 判断线上错误统计是否存在
    :param report_path_name:
    :return:
    '''
    if os.path.exists(os.path.abspath('../ReportAndLog/Statistics/Online_System_Status_Statistics.xlsx')):
        report_data_dict = read_report(report_path_name)
        file_data_dict = read_xlsx()
        merge_dict_data(report_data_dict, file_data_dict)
        write_xlsx(report_data_dict)
    else:
        report_data_dict = read_report(report_path_name)
        write_xlsx(report_data_dict)

if __name__ == '__main__':
    run_statistics_report("D:/PythonWorkSpace/HighPin_VIK/ReportAndLog/Report/result_2017-02-23_12-40-00.html")

